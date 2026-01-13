import time
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


def load_player_ids_from_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    player_ids = set()
    sheet_counts = {"RAW": 0, "Others": 0}

    for sheet_name in ("RAW", "Others"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(max_col=1, values_only=True):
            if not row[0]:
                continue

            value = str(row[0]).strip()

            if value not in player_ids:
                player_ids.add(value)
                sheet_counts[sheet_name] += 1

    wb.close()
    return list(player_ids), sheet_counts


def save_result(results_file, file_lock, player_id, attempt, status, message):
    with file_lock:
        with open(results_file, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([player_id, attempt, status, message])


def process_player(driver, wait, player_id, config, results_file, file_lock, counters, save):
    for attempt in range(1, config["MAX_RETRIES"] + 2):
        try:
            driver.get(config["URL"])

            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Enter Gift Code']")
            )).clear()
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Enter Gift Code']")
            )).send_keys(config["GIFT_CODE"])

            time.sleep(config["WAIT"])

            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Player ID']")
            )).clear()
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Player ID']")
            )).send_keys(player_id)

            time.sleep(config["WAIT"])

            wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class,'login_btn') and not(contains(@class,'disabled'))]")
            )).click()

            time.sleep(config["WAIT"])

            wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class,'exchange_btn') and not(contains(@class,'disabled'))]")
            )).click()

            popup_text = wait.until(
                EC.presence_of_element_located((By.XPATH, "//p[@class='msg']"))
            ).text

            if "Redeemed" in popup_text:
                print(f"✅ Player: {player_id} -> \"{popup_text}\", attempt: {attempt}\n", flush=True)

                with file_lock:
                    counters["redeemed"] += 1
                if save:
                    save_result(results_file, file_lock, player_id, attempt, "REDEEMED", popup_text)
                return

            if "Already claimed" in popup_text:
                print(f"✅ Player: {player_id} -> \"{popup_text}\", attempt: {attempt}\n", flush=True)

                with file_lock:
                    counters["claimed"] += 1
                if save:
                    save_result(results_file, file_lock, player_id, attempt, "ALREADY_CLAIMED", popup_text)
                return

            raise Exception(popup_text)

        except Exception as e:
            print(f"❌ Failed for Player: {player_id}, attempt: {attempt} ❌\n", flush=True)
            if save:
                save_result(results_file, file_lock, player_id, attempt, "ERROR", str(e))

            if attempt == config["MAX_RETRIES"] + 1:
                with file_lock:
                    counters["errors"].append(player_id)

            time.sleep(config["WAIT"])


def worker(player_ids, config, results_file, file_lock, counters):
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 10)

    try:
        for player_id in player_ids:
            process_player(driver, wait, player_id, config, results_file, file_lock, counters, config["SAVE_RESULTS"])
    finally:
        driver.quit()


def chunk_list(lst, chunk_size_):
    for i in range(0, len(lst), chunk_size_):
        yield lst[i:i + chunk_size_]
